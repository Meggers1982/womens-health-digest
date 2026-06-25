#!/usr/bin/env python3
"""
Extract journal data from PubMed_Journals_Categorized.xlsx into per-category CSVs.
Writes one CSV per sheet into data/, matching the column structure the digest pipeline expects.
"""

import csv
import sys
from pathlib import Path

import openpyxl

XLSX_PATH = Path("/Users/meaganmorris/PubMed_Journals_Categorized.xlsx")
OUT_DIR   = Path(__file__).parent.parent / "data"

SHEETS = [
    "Womens Health & Reproduction",
    "Endocrinology & Metabolism",
    "Psychiatry",
    "Nutritional Sciences",
]

COLUMNS = ["Journal Title", "MEDLINE Abbreviation", "ISSN (Print)", "ISSN (Online)",
           "Country", "Categories", "Start Year", "NLM ID"]


def extract_sheet(wb, sheet_name: str, out_path: Path) -> int:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print(f"  WARNING: {sheet_name} is empty")
        return 0

    header = list(rows[0])
    col_idx = {name: header.index(name) for name in COLUMNS if name in header}

    missing = [c for c in COLUMNS if c not in col_idx]
    if missing:
        print(f"  WARNING: {sheet_name} missing columns: {missing}")

    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(COLUMNS)
        written = 0
        for row in rows[1:]:
            if not any(row):
                continue
            out_row = []
            for col in COLUMNS:
                if col in col_idx:
                    val = row[col_idx[col]]
                    if val is None:
                        out_row.append("")
                    elif col == "Start Year" and isinstance(val, float):
                        out_row.append(str(int(val)))
                    else:
                        out_row.append(str(val))
                else:
                    out_row.append("")
            writer.writerow(out_row)
            written += 1
    return written


def main():
    if not XLSX_PATH.exists():
        print(f"ERROR: {XLSX_PATH} not found", file=sys.stderr)
        sys.exit(1)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    print(f"Loading {XLSX_PATH}...")
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)

    for sheet in SHEETS:
        if sheet not in wb.sheetnames:
            print(f"  ERROR: sheet '{sheet}' not found in workbook")
            continue
        out_path = OUT_DIR / f"{sheet}.csv"
        count = extract_sheet(wb, sheet, out_path)
        print(f"  {sheet}: {count} journals → {out_path.name}")

    print("Done.")


if __name__ == "__main__":
    main()
